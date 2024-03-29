# !/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Desc:
Authors: 
Date: 
"""
from time import sleep
from colorama import Fore
from arche.societies import ArcheActPlaying
from arche.utils import print_text_animated
from arche.typing import TaskType


def main(task_prompt: str):
    """
    Desc:
        main entrance of this script
    """
    session = ArcheActPlaying(
        assistant_role_name='出行助手', 
        user_role_name='游客',
        task_prompt=task_prompt,
        with_task_specify=True,
        skip_init_agents=True,
        task_specify_agent_kwargs=dict(word_limit=100),
        with_task_planner=False,
        task_type=TaskType.TRAVEL_ASSISTANT,
        extend_sys_msg_meta_dicts=None,
        extend_task_specify_meta_dict=None,
        output_language='Chinese',
    )

    if session.task_is_clarified:
        # print(
        #     Fore.GREEN +
        #     f"AI Assistant sys message:\n{session.assistant_sys_msg}\n")
        # print(Fore.BLUE +
        #     f"AI User sys message:\n{session.user_sys_msg}\n")
        pass

    print(Fore.YELLOW + f"Original task prompt:\n{task_prompt}\n")
    print(
        Fore.CYAN +
        f"Specified task prompt:\n{session.specified_task_prompt}\n")
    print(Fore.RED + f"Final task prompt:\n{session.task_prompt}\n")
    print(Fore.WHITE + "")
    return session.specified_task_prompt


def batch_test_cq_res(path2xlfile: str) -> None:
    """
    Desc:
        batch test clarified query with EB4 api
    Args:
        path2xlfile:
    Returns:
        None
    """
    import os
    from pathlib import Path
    import openpyxl

    outputPath = Path(path2xlfile)
    outputPath.parent.mkdir(parents=True, exist_ok=True)
    last_two_dirs = '_'.join(outputPath.parts[-3:-1])
    inputfilename = outputPath.stem
    dir = outputPath.resolve().parent
    output_excel_name = os.path.join(dir, f"{last_two_dirs}_{inputfilename}.xlsx")
    if Path(output_excel_name).exists():
        output_workbook = openpyxl.load_workbook(output_excel_name)
    else:
        output_workbook = openpyxl.Workbook()
    output_worksheet = output_workbook.active
    output_worksheet.cell(row=1, column=1, value='INDEX')
    output_worksheet.cell(row=1, column=2, value=f'原始 Query')
    output_worksheet.cell(row=1, column=3, value=f'Clarified Query')
    
    input_workbook = openpyxl.load_workbook(path2xlfile)
    input_sheet = input_workbook.active
    for idx, row in enumerate(input_sheet.iter_rows(min_row=2)):
        original_query = row[1].value
        output_worksheet.cell(row=idx + 2, column=1, value=idx + 1)
        output_worksheet.cell(row=idx + 2, column=2, value=original_query)
        output_worksheet.cell(row=idx + 2, column=3, value=main(original_query))
        sleep(1)
    output_workbook.save(output_excel_name)


if __name__ == '__main__':
    batch_test_cq_res('examples/travel_assistant/mix_query_20240306.xlsx')
